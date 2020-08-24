# -*- coding: utf-8 -*-

'2019-07-25 Created by zhulk'
'''
实现内容：EXCEL数据驱动，可选择执行那些测试用例，错误用例发送邮件
'''
import xlrd
from config.url import apiDir
from config.url import apiDirRes
from public.log import Log
import requests
from email.mime.multipart import MIMEMultipart
from email.header import Header
from email.mime.text import MIMEText
import smtplib
import time
import openpyxl
from public.common import NOW

log = Log()
path = apiDir
pathRes = apiDirRes

def readdata(filrpath,num):
    caseList=[]
    file = xlrd.open_workbook(filrpath)
    case = file.sheets()[0]
    nrows = case.nrows
    for i in range(num,nrows):
        caseList.append({"caseid":case.cell(i,0).value,'casename':case.cell(i,1).value,
                         'parameter':case.cell(i,2).value,'url':case.cell(i,3).value,
                         'method':case.cell(i,4).value,'expected':case.cell(i,5).value,
                         'header':case.cell(i,6).value,'correlation':case.cell(i,7).value,
                        'database':case.cell(i,8).value,'excute':case.cell(i,9)})
    return caseList

def get(url,params=None,headers=None):
    log.info('get请求开始')
    r = requests.get(url, params=params, headers=headers)
    log.info("请求的内容：%s" % params)
    statusCode = r.status_code
    log.info("获取返回的状态码:%d" % statusCode)
    try:
        responseJson = r.json()
        log.info("响应内容：%s" % responseJson)
        return statusCode,responseJson
    except Exception:
        responseText = r.text
        log.info("响应内容：%s" % responseText)
        return statusCode,responseText

def post(url,data=None,headers=None,type=None):
    log.info('post请求开始')
    if type == 'data':
        r = requests.post(url, data=data, headers=headers)
    elif type == 'json':
        r = requests.post(url, json=data, headers=headers)
    log.info("请求的内容：%s" % data)
    statusCode = r.status_code
    log.info("获取返回的状态码:%d" % statusCode)
    cookie = r.cookies
    log.info("：%s" % cookie)
    try:
        responseJson = r.json()
        log.info("响应内容：%s" % responseJson)
        return statusCode,responseJson,cookie
    except Exception:
        responseText = r.text
        log.info("响应内容：%s" % responseText)
        return statusCode,responseText,cookie

def copyExcel(path, pathRes):
    wb2 = openpyxl.Workbook()
    wb2.save(pathRes)
    wb1 = openpyxl.load_workbook(path)
    wb2 = openpyxl.load_workbook(pathRes)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row
    max_column = sheet1.max_column
    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):
            n = chr(n)
            i ='%s%d'% (n, m)
            cell = sheet1[i].value
            sheet2[i].value = cell
    wb2.save(pathRes)
    wb1.close()
    wb2.close()

def writeContent(filename,row,col,value):
    try:
        wb =openpyxl.load_workbook(filename)
        ws = wb.active
        ws.cell(row,col).value = value
        wb.save(filename)
    except Exception:
        log.info('系统错误,请尝试关闭文件')



def interfaceTest(caseList,path):
    global content
    content = ''
    passCase = 0
    failCase = 0
    detail = []
    reason = []
    i = 2
    for case in caseList:
        caseid = str(case['caseid'])
        url = str(case['url'])
        method = str(case['method'])
        expected = str(case['expected']).strip()
        header = str(case['header'])
        correlation = str(case['correlation'])
        casename = str(case['casename'])
        database = str (case['database'])
        excute = str(case['excute'])
        if 'Y' in excute.upper() :
            if case['parameter'] == '':
                parameter = ''
            else:
                parameter = eval(case['parameter'].strip())
                for parms in parameter.keys():
                    if parms == 'Token':
                        parameter['Token'] = content
                    elif parms == 'token':
                        parameter['token'] = content
            if method.upper() == 'GET':
                if parameter == '' and header == '':
                    res = get(url)
                elif header == '':
                    res = get(url,params = parameter)
                elif parameter == '':
                    res = get(url,headers = header)
                else:
                    res = get(url,params = parameter,headers = header)
                if expected in str(res[1]) and res[0] == 200:
                    detail.append('编号' + caseid + casename + '测试通过')
                    writeContent(path, i, 10, res[0])
                    writeContent(path, i, 11, str(res[1]))
                    writeContent(path, i, 12, '测试通过')
                    passCase = passCase + 1
                else:
                    detail.append('编号' + caseid + casename + '测试失败')
                    writeContent(path, i, 10, res[0])
                    writeContent(path, i, 11, str(res[1]))
                    writeContent(path, i, 12, '测试失败')
                    reason.append(caseid + casename + '返回值：' + str(res[1]) + "&nbsp" * 6 +'断言值：' + expected)
                    failCase = failCase + 1
            if method.upper() == 'POST':
                if database =='data':
                    if parameter == '' and header == '':
                        res = post(url,type = 'data')
                    elif header == '':
                        res = post(url,data = parameter,type = 'data')
                    elif parameter == '':
                        res = post(url,headers=header,type = 'data')
                    else:
                        res = post(url,data = parameter,headers = header,type = 'data')
                    if expected in str(res[1]) and res[0] == 200:
                        detail.append('编号'+caseid+casename+'测试通过')
                        passCase = passCase + 1
                        writeContent(path,i,10, res[0])
                        writeContent(path,i,11, str(res[1]))
                        writeContent(path,i,12,'测试通过')
                    else:
                        detail.append('编号'+caseid+casename+'测试失败')
                        writeContent(path, i, 10, res[0])
                        writeContent(path, i, 11, str(res[1]))
                        writeContent(path, i, 12, '测试失败')
                        reason.append(caseid+casename+'返回值：'+ str(res[1])+ "&nbsp" * 6 + '断言值：'+expected)
                        failCase = failCase + 1
                    if correlation == 'token':
                        content = res[2]['tn']
                elif database =='json':
                    if parameter == '' and header == '':
                        res = post(url,type = 'json')
                    elif header == '':
                        res = post(url,data = parameter,type = 'json')
                    elif parameter == '':
                        res = post(url,headers=header,type = 'json')
                    else:
                        res = post(url,data = parameter,headers = header,type = 'json')
                    if expected in str(res[1]):
                        detail.append('编号'+caseid+casename+'测试通过')
                        writeContent(path, i, 10, res[0])
                        writeContent(path, i, 11, str(res[1]))
                        writeContent(path, i, 12, '测试通过')
                        passCase = passCase + 1
                    else:
                        detail.append('编号'+caseid+casename+'测试失败')
                        writeContent(path, i, 10, res[0])
                        writeContent(path, i, 11, str(res[1]))
                        writeContent(path, i, 12, '测试失败')
                        reason.append(caseid+casename+'返回值：'+ str(res[1])+ "&nbsp" * 6 + '断言值：'+expected)
                        failCase = failCase + 1
                    # if correlation == 'token':
                    #     content = res[2]['tn']
        elif 'N' in excute.upper():
            writeContent(path, i, 10, '')
            writeContent(path, i, 11, '')
            writeContent(path, i, 12, '未执行')
        i = i + 1
    if reason == []:
        res = ('成功' + str(passCase) + '个用例,失败' + str(failCase) + '个用例<br />') + str(
                        detail)
    else:
        res = ('成功'+ str(passCase)+'个用例,失败'+str(failCase)+'个用例<br />')+str(detail)+'<br />'+str(reason)
    if failCase != 0:
        sendWarmEmail(res)
    writeContent(path, 3, 15, '成功用例')
    writeContent(path, 3, 16, passCase)
    writeContent(path, 4, 15, '失败用例')
    writeContent(path, 4, 16, failCase)
    return res

def sendWarmEmail(content):
    mailfrom = 'zlk_0102@163.com'
    mailto = '517122472@qq.com'
    msg = MIMEMultipart()
    msg.attach(MIMEText(content, _subtype='html', _charset='utf-8'))
    msg['Subject'] = Header(u'商旅测试接口自动化测试报告', 'utf-8')
    msg['date'] = time.strftime('%a, %d %b %Y %H:%M:%S %z')
    msg['From'] = mailfrom
    msg['to'] = mailto
    att = MIMEText(open(pathRes, 'rb').read(), 'base64', 'utf-8')
    att["Content-Type"] = 'application/octet-stream'
    att["Content-Disposition"] = 'attachment; filename=' + pathRes
    msg.attach(att)
    username = 'zlk_0102@163.com'
    password = 'TLNOFJSYQPYWISOA'
    smtp = smtplib.SMTP()
    smtp.connect('smtp.163.com')
    smtp.login(username, password)
    smtp.sendmail(mailfrom, mailto, msg.as_string())
    smtp.quit()

copyExcel(path,pathRes)
writeContent(pathRes,1,10,'状态码')
writeContent(pathRes,1,11,'响应内容')
writeContent(pathRes,1,12,'测试结果')
writeContent(pathRes,1,13,'运行完成时间')
writeContent(pathRes,2,13,NOW())
interfaceTest(readdata(pathRes,1),pathRes)
